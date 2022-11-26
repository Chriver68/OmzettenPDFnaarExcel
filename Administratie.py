import streamlit as st
import pdfplumber
from openpyxl import Workbook, load_workbook
from datetime import datetime

rekening_file = ''

def wegschrijven():
    #try:
    wb = load_workbook(bestand)
    ws = wb.active
    ws.title = 'Facturen'
    now = datetime.now()
    schrijfmoment = now.strftime("%d-%m-%Y %H:%M:%S")
    ws.append([str(factuurnummer[0]), factuurdatum_conv, float(subtotaal[0]), float(btw_bedrag[0])
                  ,float(totaal[0]), opdrachtgever, rekening_file, schrijfmoment])
    wb.save(bestand)
    st.info(f'Gegevens zijn toegevoegd aan het excel overzicht {bestand}!')
    #except:
        #st.error('Er is iets fout gegaan! Was het bestand nog geopend? Sluit deze eerst en probeer het dan nog een keer!')

def controle():
    #try:
    wb = load_workbook(bestand)
    ws = wb.active
    ws.title = 'Facturen'
    bestaat_al = ''
    for data in ws["G"]:
        if data.value == rekening.name:
            bestaat_al = 'Ja'

    wb.close()
    if bestaat_al != 'Ja':
        wegschrijven()
    else:
        st.warning('Het bestand was al verwerkt in het Excel overzicht!')
    #except:
        #st.error('Er is iets fout gegaan! Was het bestand nog geopend? Sluit deze eerst en probeer het dan nog een keer!')


st.header('Omzetten factuurinformatie naar Excel')
st.text(f"""Factuurnummer: Factuurdatum: Sub-totaal: BTW Bedrag: Totaal: Opdrachtgever: Bestand:
""")
st.subheader('Selecteer hier de betreffende factuur:')
rekening = st.file_uploader(('Selecteer een PDF factuur om te converteren naar CSV/Excel'))

if rekening is not None:
    with pdfplumber.open(rekening) as pdf:
        page = pdf.pages[0]
        text = page.extract_text()

    bestand_zien = st.radio('Wil je het bestand zien?', ('Nee', 'Ja'), horizontal=True, label_visibility= 'visible')
    if bestand_zien == 'Ja':
        st.write(text)

    words = str.split(text)
    #aantal = 0
    x=0

    for word in words:
        x += 1
        if word == 'Factuurdatum:':
            factuurdatum = words[x:x+1]

        elif word == 'Factuurnummer:':
            factuurnummer = words[x:x + 1]

        elif word == '(excl'and words[x] == 'btw)':
            subtotaal = words[x+2:x + 3]

        elif word == 'BTW':
           btw_bedrag = words[x+1:x+2]

        elif word == '(incl'and words[x] == 'btw)':
            totaal = words[x+2:x + 3]

    st.subheader(f'Controleer de volgende gegevens:')
    st.write(f'Bestandsnaam {rekening.name}')
    st.write(f'Factuurnummer: {factuurnummer[0]}')
    st.write(f'Factuurdatum: {factuurdatum[0]}')
    st.write(f'Sub-totaal: {subtotaal[0]}')
    st.write(f'BTW Bedrag: {btw_bedrag[0]}')
    st.write(f'Totaal: {totaal[0]}')

    rekening_file = str(rekening.name)
    start= rekening_file.find('wa-')+3
    eind = rekening_file.find('-Factuur')
    opdrachtgever = rekening_file[start:eind]
    st.write(f'Opdrachtgever: {opdrachtgever}')
    factuurdatum_conv = str(factuurdatum[0])

bestand = st.text_input(label = "Geef het pad op waar het Excel bestand staat waar de gegevens in moeten.", label_visibility= 'visible' )
if bestand != '' and rekening_file != '':
    if(st.button('Akkoord en opslaan in Excel!')):
        controle()

    
